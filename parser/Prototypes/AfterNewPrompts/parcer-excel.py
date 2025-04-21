import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re
import os

file_path = os.path.join(os.getcwd(), "Parcer", "TestingSpreadSheet.xlsx")
book = load_workbook(file_path)

def extract_rankings_claude(file_path):
    with open(file_path, "r") as file:
        content = file.read()
    
    content = content.replace("\r\n", "\n").replace("\r", "\n")
    
    match = re.search(r":\\n\\n((?:\d+,\s*\d+\\n?)+)", content)
    if match:
        extracted_text = match.group(1).strip()

        patients = []
        rankings = []
        
        for line in extracted_text.split("\\n"):
            line = line.strip()
            parts = line.split(",")
            
            if len(parts) == 2:
                try:
                    patient = int(parts[0].strip()) 
                    rank = int(parts[1].strip())
                    
                    patients.append(patient)
                    rankings.append(rank)
                except ValueError:
                    print(f"Skipping invalid line: {repr(line)}")
        
        sorted_patients, sorted_rankings = zip(*sorted(zip(patients, rankings)))
        min_patient = sorted_patients[0]
        max_patient = sorted_patients[-1]
        
        full_patients = list(range(min_patient, max_patient + 1))
        full_rankings = [0] * (max_patient - min_patient + 1)
        
        for patient, rank in zip(sorted_patients, sorted_rankings):
            full_rankings[patient - min_patient] = rank
        
        return full_patients, full_rankings
        
    
    else:
        print(f"No matching, kidney {currKidney} ,shuffle {currShuffle}, trial {currTrial}")
        return [], []

def extract_rankings_GPT(file_path):
    with open(file_path, "r") as file:
        content = file.read()
    
    content = content.replace("\r\n", "\n").replace("\r", "\n")
    
    match = re.search(r":\n\n((?:(\d+),\s*(\d+)\n?)+)", content)
    if match:
        extracted_text = match.group(1).strip()

        patients = []
        rankings = []
        
        for line in extracted_text.split("\\n"):
            line = line.strip()
            parts = line.split(",")
            
            if len(parts) == 2:
                try:
                    patient = int(parts[0].strip()) 
                    rank = int(parts[1].strip())
                    
                    patients.append(patient)
                    rankings.append(rank)
                except ValueError:
                    print(f"Skipping invalid line: {repr(line)}")
        
        sorted_patients, sorted_rankings = zip(*sorted(zip(patients, rankings)))
        min_patient = sorted_patients[0]
        max_patient = sorted_patients[-1]
        
        full_patients = list(range(min_patient, max_patient + 1))
        full_rankings = [0] * (max_patient - min_patient + 1)
        
        for patient, rank in zip(sorted_patients, sorted_rankings):
            full_rankings[patient - min_patient] = rank
        
        return full_patients, full_rankings
        
    
    else:
        print(f"No matching, kidney {currKidney} ,shuffle {currShuffle}, trial {currTrial}")
        return [], []

types = ["all", "demo_only"]
bok = ["Claude-All", "Claude-Demo"]
index = 0
for index in range(2):
    sheet = book[bok[index]]
    for currKidney in range(20):
        for currShuffle in range(3):
            col_index = 3 + 8*currShuffle 
                    
            for currTrial in range(5):
                file_path = f"results/claude/ranked/ranked_kidney{currKidney}_{types[index]}_shuffle{currShuffle}_trial{currTrial}.out"
                patients, rankings = extract_rankings_claude(file_path)

                if patients and rankings:

                    column_letter = get_column_letter(col_index)
                    for i, rank in enumerate(rankings, start=4+(currKidney*17)):
                        sheet[f"{column_letter}{i}"] = rank

                    col_index += 1
                            
                    book.save('TestingSpreadSheet.xlsx')
        print(f'Kidney {currKidney} updated')
    print(f'Type: {types[index]} updated')

