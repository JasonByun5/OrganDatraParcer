import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re
import json

book = load_workbook('TestingSpreadSheet.xlsx')
sheet = book.active

def extract_patient_rankings(file_path):
    rankings = []
    with open(file_path, "r") as file:
        for line in file:
            try:
                data = json.loads(line)
                if "content" in data and isinstance(data["content"], str):
                    ranking_lines = data["content"].strip().split("\n")
                    for ranking in ranking_lines:
                        match = re.match(r"(\d+),\s*(\d+)", ranking.strip())
                        if match:
                            patient_id, rank = int(match.group(1)), int(match.group(2))
                            rankings.append((patient_id, rank))
            except json.JSONDecodeError:
                continue  # Skip lines that are not valid JSON

    # Sort rankings based on the patient ID for consistency
    rankings.sort(key=lambda x: x[0])
    
    if not rankings:
        print(f"No valid ranking data found in {file_path}.")
    
    return rankings



currShuffle = 0
currTrial = 0
currKidney = 0


#for currShuffle in range(2):
col_index = 2 + 8*currShuffle 
        
#for currTrial in range(5):


file_path = f"ranked-claude/ranked_kidney{currKidney}_all_shuffle{currShuffle}_trial{currTrial}.out"
rankings = extract_patient_rankings(file_path)



column_letter = get_column_letter(col_index)
for i, r in enumerate(rankings, start = 3):
    sheet[f'{column_letter}{i}'] = r

col_index += 1
        
book.save('TestingSpreadSheet.xlsx')
print(f'Shuffle {currShuffle} updated')

