import openpyxl
from openpyxl.utils import get_column_letter
import json
import re

# File paths
file_paths = [
    "ranked-claude/ranked_kidney0_all_shuffle0_trial0.out",
    "ranked-claude/ranked_kidney0_all_shuffle0_trial1.out",
    "ranked-claude/ranked_kidney0_all_shuffle0_trial2.out"
]

# Load the Excel workbook
excel_path = "/mnt/data/TestingSpreadSheet.xlsx"
try:
    book = openpyxl.load_workbook(excel_path)
except FileNotFoundError:
    book = openpyxl.Workbook()

sheet = book.active

def extract_rankings_from_json(file_path):
    """Extract ranking data from JSON-formatted text."""
    rankings = []
    with open(file_path, "r") as file:
        for line in file:
            try:
                data = json.loads(line)
                if "content" in data and isinstance(data["content"], str):
                    ranking_lines = data["content"].strip().split("\n")
                    for ranking in ranking_lines:
                        match = re.match(r"(\d+),\s*(\d+)", ranking.strip())
                        if match:
                            patient_id, rank = int(match.group(1)), int(match.group(2))
                            rankings.append((patient_id, rank))
            except json.JSONDecodeError:
                continue  # Skip lines that are not valid JSON

    # Sort rankings based on the patient ID for consistency
    rankings.sort(key=lambda x: x[0])
    
    if not rankings:
        print(f"No valid ranking data found in {file_path}.")
    
    return rankings

# Initialize column index based on shuffle/trial assumptions
currShuffle = 0
col_index = 2 + 8 * currShuffle  

# Extract rankings from all files
all_rankings = []
for file_path in file_paths:
    rankings = extract_rankings_from_json(file_path)
    all_rankings.append(rankings)

# Write extracted rankings to the Excel sheet
for currTrial, rankings in enumerate(all_rankings):
    column_letter = get_column_letter(col_index + currTrial)
    
    # Write rankings to the Excel sheet
    for i, (patient_id, rank) in enumerate(rankings, start=3):
        sheet[f"{column_letter}{i}"] = rank

# Save the updated Excel file
updated_excel_path = "/mnt/data/Final_Updated_TestingSpreadSheet.xlsx"
book.save(updated_excel_path)

print(f"Updated spreadsheet saved at: {updated_excel_path}")