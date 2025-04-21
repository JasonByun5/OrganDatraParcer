import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re

book = load_workbook('TestingSpreadSheet.xlsx')
sheet = book.active

def extract_patient_rankings(file_path):
    patient_dict = {}
    
    with open(file_path, "r") as file:
        for line in file:
            match = re.match(r"(\d+),(\d+)", line.strip())
            if match:
                rank, patient = (int, match.group())
                if patient in patient_dict:
                    print(f"Duplicate entry found for patient {patient}, keeping the first occurrence.")
                else:
                    patient_dict[patient] = rank
    
    if not patient_dict:
        print("No, valid patient ranking data found in the file.")
        return [], []
            
    sorted_patients = sorted(patient_dict.keys())
    sorted_rankings = [patient_dict[p] for p in sorted_patients]
    
    max_patient_id = max(sorted_patients, default = 0)
    full_patients = list(range(1, max_patient_id + 1))
    full_rankings = [patient_dict.get(p, "None") for p in full_patients]
    
    return full_patients, full_rankings



currShuffle = 0
currTrial = 0
currKidney = 0


#for currShuffle in range(2):
col_index = 2 + 8*currShuffle 
        
#for currTrial in range(5):


file_path = f"ranked-claude/ranked_kidney{currKidney}_all_shuffle{currShuffle}_trial{currTrial}.out"
patients, rankings = extract_patient_rankings(file_path)
print("Sorted Patients:", patients)
print("Sorted Rankings:", rankings)


column_letter = get_column_letter(col_index)
for i, r in enumerate(rankings, start = 3):
    sheet[f'{column_letter}{i}'] = r

col_index += 1
        
book.save('TestingSpreadSheet.xlsx')
print(f'Shuffle {currShuffle} updated')

