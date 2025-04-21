import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re
import os

file_path = os.path.join(os.getcwd(), "Parcer", "TestingSpreadSheet.xlsx")
book = load_workbook(file_path)


# Open the file and read its contents
def print_files(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        content = file.read()

    # Print the content
    print(content)

def extract_rankings_GPT(file_path):
    with open(file_path, "r") as file:
        content = file.read()
    
    content = content.replace("\r\n", "\n").replace("\r", "\n")
    content = content.replace("\\n", "\n")

    #Case 1: match = re.search(r":\n\n```[\n\r]*((?:\d+,\s*\d+\s*\n?)+)", content)
    #Case 2: match = re.search(r":\n\n((?:\d+,\s*\d+\s*\n?)+)", content)

    #Combined!!
    match = re.search(r":\n\n(?:```)?[\n\r]*((?:\d+,\s*\d+\s*\n?)+)", content)
    
    if match:
        extracted_text = match.group(1).strip()
        
        print("extracted_text: ", extracted_text)

        patients = []
        rankings = []
        
        for line in extracted_text.split("\n"):
            line = line.strip()
            parts = line.split(",")
            
            if len(parts) == 2:
                try:
                    patient = int(parts[0].strip()) 
                    rank = int(parts[1].strip())
                    
                    patients.append(patient)
                    rankings.append(rank)
                except ValueError:
                    print(f"Skipping invalid line: {repr(line)}")
        
        print("Pre sorted Patients: ", patients)
        print("Pre sorted Rankings: ", rankings)
        sorted_patients, sorted_rankings = zip(*sorted(zip(patients, rankings)))
        min_patient = sorted_patients[0]
        max_patient = sorted_patients[-1]
        
        full_patients = list(range(min_patient, max_patient + 1))
        full_rankings = [0] * (max_patient - min_patient + 1)
        
        for patient, rank in zip(sorted_patients, sorted_rankings):
            full_rankings[patient - min_patient] = rank
        
        return full_patients, full_rankings
    
    else:
        print(f"No matching, kidney {currKidney} ,shuffle {currShuffle}, trial {currTrial}")
        return [], []
 
types = ["all", "demo_only"]
currKidney = 0
index = 0
currShuffle = 0
currTrial = 0

file_path = f"results/gpt/ranked/ranked_kidney{currKidney}_{types[index]}_shuffle{currShuffle}_trial{currTrial}.out"
print_files(file_path)
patients, rankings = extract_rankings_GPT(file_path)

print("Found Patients: ", patients)
print("Found Rankings: ", rankings)

