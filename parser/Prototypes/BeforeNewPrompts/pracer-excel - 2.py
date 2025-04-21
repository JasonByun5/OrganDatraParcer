import re
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

book = load_workbook('TestingSpreadSheet.xlsx')
sheet = book.active

def extract_patient_rankings(file_path):
    patients = []
    rankings = []
    
    with open(file_path, "r") as file:
        for line in file:
            match = re.match(r'^(\d+),\s*(\d+),', line.strip())
            if match:
                patient_id = match.group(1)
                ranking = int(match.group(2))
                patients.append(patient_id)
                rankings.append(ranking)
    
    if not patients or not rankings:
        print("No valid patient ranking data found in the file.")
        return [], []
    
    sorted_patients, sorted_rankings = zip(*sorted(zip(patients, rankings)))
    sorted_rankings = list(sorted_rankings)
    
    return list(sorted_patients), sorted_rankings



currShuffle = 0
currTrial = 2
for currShuffle in range(2):
    col_index = 2 + 8*currShuffle 
        
    for currTrial in range(5):

        file_path = f'ranked_kidney0_all_shuffle{currShuffle}_trial{currTrial}.out'
        patients, rankings = extract_patient_rankings(file_path)
        print("Patients:", patients)
        print("Rankings:", rankings)


        column_letter = get_column_letter(col_index)

        for i, r in enumerate(rankings, start = 3):
            sheet[f'{column_letter}{i}'] = r

        col_index += 1
                
    book.save('TestingSpreadSheet.xlsx')
    print(f'Shuffle {currShuffle} updated')

