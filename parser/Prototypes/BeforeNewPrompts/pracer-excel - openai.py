import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openai
import ast
# import os
# from dotenv import load_dotenv

# api_key = os.getenv("OPEN_API_KEY")
openai_client = openai.Client(api_key="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") 

book = load_workbook('TestingSpreadSheet.xlsx')
sheet = book.active


currShuffle = 0
currTrial = 0
#for currShuffle in range(2):
col_index = 2 + 8 * currShuffle 


for currTrial in range(5):
    with open (f'ranked_kidney0_all_shuffle{currShuffle}_trial{currTrial}.out', 'r') as f:
        text_data = f.read()
        
    if text_data:
        messages = [
            {"role": "user", "content": f"If I was to give you this file {text_data}, past line 80 there are the rankings that I need you to parse. The first value in the line is the patient as the second value is the ranking. Output ONLY the 2 arrays and do not give labels to them"}
        ]

        response = openai_client.chat.completions.create(  # Updated method call
            model="gpt-4o-mini",
            messages=messages
        )

        gpt_response = response.choices[0].message.content

        lines = gpt_response.strip().split("\n")

        rankings = ast.literal_eval(lines[0])
        patients = ast.literal_eval(lines[1])

        print(patients)
        print(rankings)


        #sorts the lists together
        sorted_patients, sorted_ranks = zip(*sorted(zip(patients, rankings)))
        sorted_ranks = list(sorted_ranks)


        column_letter = get_column_letter(col_index)

        for i, r in enumerate(sorted_ranks, start = 3):
            sheet[f'{column_letter}{i}'] = r

        col_index += 1
    print(f'{currTrial} Updated')         

book.save('TestingSpreadSheet.xlsx')
print(f'Shuffle {currShuffle} updated')



