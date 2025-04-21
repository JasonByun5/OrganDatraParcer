
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

book = load_workbook('TestingSpreadSheet.xlsx')
sheet = book.active


currShuffle = 1
currTrial = 0
#for currShuffle in range(2):
col_index = 2 + 8*currShuffle 
    
#for currTrial in range(5):
with open (f'ranked_kidney0_all_shuffle{currShuffle}_trial{currTrial}.out', 'r') as f:
    for line in f:
        line = line.strip()
        if ("Here's the ranking" in line) or ("I'll rank the 15 candidates based" in line) or ("rankings" in line):
            print(line)
            print("Match found")
            break
        
    line = f.readline().strip() #used to get the null string
    
    patients = []
    ranks = [] # when sorting the patients list, what you change in the patient list, change in the rank list
    
    for i in range(15):
        line = f.readline().strip()
        
        temp = line.split(",")
        
        patients.append(int(temp[0]))
        ranks.append(int(temp[1].strip()))
    
    #sorts the lists together
    sorted_patients, sorted_ranks = zip(*sorted(zip(patients, ranks)))
    sorted_ranks = list(sorted_ranks)
    
    
    column_letter = get_column_letter(col_index)
    
    for i, r in enumerate(sorted_ranks, start = 3):
        sheet[f'{column_letter}{i}'] = r
    
    col_index += 1
        
book.save('TestingSpreadSheet.xlsx')
print(f'Shuffle {currShuffle} updated')


